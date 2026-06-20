
import openpyxl
import os
import random
accounts = []

def setup_excel():
    """Creates the Excel file if it doesn't exist yet"""
    if not os.path.exists("lena_dena_bank.xlsx"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Accounts"
        # These are the column headers
        ws.append(["Account Number", "Name", "Age", "Phone", "City", "Balance"])
        wb.save("lena_dena_bank.xlsx")
        print("Bank file created!")

def load_accounts():
    """Load existing accounts from Excel into the accounts list when app starts"""
    if not os.path.exists("lena_dena_bank.xlsx"):
        return
    wb = openpyxl.load_workbook("lena_dena_bank.xlsx")
    ws = wb["Accounts"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:      # skip empty rows
            customer = {
                "acc_number" : row[0],
                "name"       : row[1],
                "age"        : row[2],
                "phone"      : row[3],
                "city"       : row[4],
                "balance"    : row[5]
            }
            accounts.append(customer)

def show_menu():
    print("\n==============================")
    print("   Welcome to Lena Dena Bank ")
    print("\n==============================")
    print("1.     Add New Account")
    print("2.      Deposit Money")
    print("3.       Withdraw Money")
    print("4.       Check Balance")
    print("5.   Search Account By Name")
    print("6.       Transfer Money")
    print("7.           Exit")
    print("\n==============================")

def add_account():
    print ("We are now in add_account function")
    customer_name = input("Enter Customer Name             :")
    customer_age = input("Enter Customer Age               :")
    customer_city = input("Enter Customer City             :")
    customer_phone = input("Enter Customer Phone           :")
    customer_balance = input("Enter Initial Deposit Amount :")

    # ── VALIDATE NAME ──────────────────────────────────────
    if not customer_name.replace(" ", "").isalpha():
        print("⚠️ Name should only have letters, no numbers or symbols!")
        return

    # ── VALIDATE AGE ───────────────────────────────────────
    if not customer_age.isdigit():
        print("⚠️ Age should only be a number!")
        return
    customer_age = int(customer_age)       # convert AFTER checking
    if customer_age < 1 or customer_age > 100:
        print("⚠️ Age must be between 1 and 100!")
        return

    # ── VALIDATE PHONE ─────────────────────────────────────
    if not customer_phone.isdigit():
        print("⚠️ Phone number should only have digits, no spaces, alphabets or dashes!")
        return
    if len(customer_phone) != 10:
        print("⚠️ Phone number must be exactly 10 digits!")
        return

    # ── VALIDATE CITY ──────────────────────────────────────
    if not customer_city.replace(" ", "").isalpha():
        print("⚠️ City name should only have letters!")
        return

    # ── VALIDATE BALANCE ───────────────────────────────────
    if not customer_balance.isdigit():
        print("⚠️ Balance should only be a number!")
        return
    customer_balance = float(customer_balance)    # convert AFTER checking
    if customer_balance < 0:
        print("⚠️ Opening balance cannot be negative!")
        return

    # ── ALL VALIDATIONS PASSED — continue with rest of function ──
    print("✅ All details look good!")
    # ... rest of your add_account code here ...


                # ✅ REPLACE WITH THIS
    if len(accounts) == 0:
        acc_number = 100001          # first ever account starts at 100001
    else:
        largest = max(customer["acc_number"] for customer in accounts)
        acc_number = largest + 1     # always 1 greater than the largest

    customer = {
        "acc_number"    : acc_number,
        "name"          : customer_name,
        "age"           : customer_age,
        "phone"         : customer_phone,
        "city"          : customer_city,
        "balance"       : customer_balance
        
    }

    accounts.append(customer)

    # Save to Excel
    wb = openpyxl.load_workbook("lena_dena_bank.xlsx")
    ws = wb["Accounts"]
    ws.append([acc_number, customer_name, customer_age, customer_phone, customer_city, float(customer_balance)])
    wb.save("lena_dena_bank.xlsx")

    print(f"\n✅ Account created! Your Account Number is: {acc_number}")

def update_excel():
    """Rewrites the whole Excel file with updated data"""
    wb = openpyxl.load_workbook("lena_dena_bank.xlsx")
    ws = wb["Accounts"]

    # ✅ Properly DELETE all rows except the header
    ws.delete_rows(2, ws.max_row)

    # Now write fresh data cleanly
    for customer in accounts:
        ws.append([
            customer["acc_number"],
            customer["name"],
            customer["age"],
            customer["phone"],
            customer["city"],
            customer["balance"]
        ])
    wb.save("lena_dena_bank.xlsx")
    print("💾 Excel file updated!")

def deposit():
    print("\n--- Deposit Money ---")

    try:
        acc_number = int(input("Enter Account Number : "))
        deposit     = float(input("Enter deposit amount : "))
    except ValueError:
        print("⚠️ Please enter valid numbers only!")
        return

    if deposit <= 0:
        print("⚠️ Deposit amount must be greater than zero!")
        return

    # Search for the account
    found = False
    for customer in accounts:
        if customer["acc_number"] == acc_number:
            old_balance          = customer["balance"]
            customer["balance"] += deposit
            update_excel()
            print(f"\n✅ Deposit successful!")
            print(f"   Account  : {acc_number}")
            print(f"   Name     : {customer['name']}")
            print(f"   Deposited: ₹{deposit}")
            print(f"   Old Bal  : ₹{old_balance}")
            print(f"   New Bal  : ₹{customer['balance']}")
            found = True
            break

    if not found:
        print("❌ Account not found. Please check the account number.")

def withdrawal():
    print("\n--- Withdraw Money ---")

    try:
        acc_number = int(input("Enter Account Number : "))
        withdrawal     = float(input("Enter withdrawal amount : "))
    except ValueError:
        print("⚠️ Please enter valid numbers only!")
        return

    if withdrawal <= 0:
        print("⚠️ Withdrawal amount must be greater than zero!")
        return
    
    # Search for the account
    found = False
    for customer in accounts:
        if customer["acc_number"] == acc_number:
            old_balance          = customer["balance"]
            if withdrawal > customer["balance"]:
                print(f"❌ Not enough balance! You only have ₹{customer['balance']}")

            else:
                customer["balance"] -= withdrawal
                update_excel()
            print(f"\n✅ Withdrawal successful!")
            print(f"   Account  : {acc_number}")
            print(f"   Name     : {customer['name']}")
            print(f"   Withdrawn: ₹{withdrawal}")
            print(f"   Old Bal  : ₹{old_balance}")
            print(f"   New Bal  : ₹{customer['balance']}")
            found = True
            break

    if not found:
        print("❌ Account not found. Please check the account number.")

def check_balance():
    print("\n--- Check Balance ---")

    try:
        acc_number = int(input("Enter Account Number : "))

    except ValueError:
        print("⚠️ Please enter valid numbers only!")
        return
    
    found = False
    for customer in accounts:
        if customer["acc_number"] == acc_number:
            print(f"\n👤 Name    : {customer['name']}")
            print(f"🏙️  City    : {customer['city']}")
            print(f"💰 Balance : ₹{customer['balance']}")
            found = True
            break

    if not found:
        print("❌ Account not found.")

def search_by_name():
    print("\n--- Search Account by Name ---")

    search_term = input("Enter name to search : ").strip()

    # Validate — don't search if input is empty
    if search_term == "":
        print("⚠️ Please type something to search!")
        return

    # Go through every account and look for a match
    results = []                         # empty list to store matches

    for customer in accounts:
        # .lower() makes search work whether you type "AR" or "ar" or "Ar"
        if search_term.lower() in customer["name"].lower():
            results.append(customer)     # found a match! add to results

    # Now show what we found
    if len(results) == 0:
        print(f"❌ No accounts found with '{search_term}' in the name.")
    else:
        print(f"\n✅ Found {len(results)} account(s):\n")
        print(f"{'AccNo':<10} {'Name':<20} {'City':<15} {'Balance'}")
        print("-" * 55)
        for customer in results:
            print(f"{customer['acc_number']:<10} {customer['name']:<20} {customer['city']:<15} ₹{customer['balance']}") 

def transfer_money():
    print("\n--- Transfer Money ---")

    try:
        from_acc = int(input("Transfer FROM account number : "))
        to_acc   = int(input("Transfer TO account number   : "))
        amount   = float(input("Amount to transfer           : "))
    except ValueError:
        print("⚠️ Please enter valid numbers only!")
        return

    # ── CHECK 1: Can't transfer to yourself! ──────────────
    if from_acc == to_acc:
        print("⚠️ You can't transfer money to the same account!")
        return

    # ── CHECK 2: Amount must be positive ──────────────────
    if amount <= 0:
        print("⚠️ Transfer amount must be greater than zero!")
        return

    # ── FIND BOTH ACCOUNTS ────────────────────────────────
    sender   = None
    receiver = None

    for customer in accounts:
        if customer["acc_number"] == from_acc:
            sender = customer         # found the sender!
        if customer["acc_number"] == to_acc:
            receiver = customer       # found the receiver!

    # ── CHECK 3: Does sender account exist? ───────────────
    if sender is None:
        print(f"❌ Sender account {from_acc} not found!")
        return

    # ── CHECK 4: Does receiver account exist? ─────────────
    if receiver is None:
        print(f"❌ Receiver account {to_acc} not found!")
        return

    # ── CHECK 5: Does sender have enough money? ───────────
    if amount > sender["balance"]:
        print(f"❌ Not enough balance!")
        print(f"   {sender['name']} only has ₹{sender['balance']}")
        return

    # ── ALL CHECKS PASSED — DO THE TRANSFER! ─────────────
    sender["balance"]   -= amount      # take money from sender
    receiver["balance"] += amount      # give money to receiver

    update_excel()                     # save both changes to Excel

    # ── PRINT RECEIPT ─────────────────────────────────────
    print(f"\n✅ Transfer Successful!")
    print(f"{'─' * 40}")
    print(f"  FROM : {sender['name']} ({from_acc})")
    print(f"  TO   : {receiver['name']} ({to_acc})")
    print(f"  AMOUNT    : ₹{amount}")
    print(f"{'─' * 40}")
    print(f"  {sender['name']}'s new balance   : ₹{sender['balance']}")
    print(f"  {receiver['name']}'s new balance : ₹{receiver['balance']}")                  

setup_excel()
load_accounts()

while True:
    show_menu()
    try:
        choice = int(input("Enter your choice(1-7): "))
    except ValueError:
        print("Please enter a number between 1-7, no characters or letters, please")
        continue

    if choice == 1:
        print("You chose: Add Account")
#       print ("Call Add Module")
        add_account()
    elif choice == 2:
        print("You chose: Deposit")
        deposit()
#       print("This function is not ready yet, try again later")
    elif choice == 3:
        print("You chose: Withdraw")
        withdrawal()
    elif choice == 4:
        print("You chose: Check Balance")
        check_balance()
    elif choice == 5:
        print("You chose: Search Account By Name")
        search_by_name()
    elif choice == 6:
        print("You chose: Transfer Money")
        transfer_money()    
    elif choice == 7:
        print("Thank you for using Lena Dena Bank. Goodbye!")
        break
    else:
        print("Oops! That's not a valid choice. Try again.")